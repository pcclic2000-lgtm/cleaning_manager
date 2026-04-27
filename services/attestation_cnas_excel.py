# services/attestation_cnas_excel.py
import os
from datetime import datetime
from pathlib import Path
from typing import Optional
from decimal import Decimal
from PyQt6.QtCore import QObject, pyqtSignal

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database.db import SessionLocal, get_session
from models.employee import Employee
from models.payslip import Payslip
from models.company import CompanyInfo


class AttestationCNASExcelGenerator(QObject):
    """
    Générateur d'attestation CNAS AS-08 au format Excel
    Exactement comme le formulaire officiel
    """
    
    progress_updated = pyqtSignal(int, str)
    generation_finished = pyqtSignal(str, bool)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
    
    def _format_date(self, date_value) -> str:
        """Formate une date au format JJ/MM/AAAA"""
        if not date_value:
            return ""
        if isinstance(date_value, datetime):
            return date_value.strftime('%d/%m/%Y')
        if isinstance(date_value, str):
            try:
                dt = datetime.strptime(date_value, '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
            except:
                return date_value
        return str(date_value)
    
    def _to_float(self, value) -> float:
        """Convertit en float"""
        if value is None:
            return 0.0
        if isinstance(value, (float, int)):
            return float(value)
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(',', '.'))
            except:
                return 0.0
        return 0.0
    
    def _format_currency(self, value) -> str:
        """Formate un montant en DA"""
        try:
            val = self._to_float(value)
            if val == 0:
                return ""
            return f"{val:,.0f}".replace(',', ' ')
        except:
            return ""
    
    def generate_for_employee(self, employee_id: int, 
                            output_path: Optional[str] = None) -> str:
        """
        Génère l'attestation AS-08 au format Excel
        """
        self.progress_updated.emit(10, "Chargement des données...")
        
        try:
            with get_session() as session:
                # ----- 1. CHARGER L'EMPLOYÉ -----
                employee = session.query(Employee).filter(
                    Employee.id == employee_id,
                    Employee.est_actif == True
                ).first()
                
                if not employee:
                    raise ValueError(f"Employé ID {employee_id} introuvable ou inactif")
                
                self.progress_updated.emit(20, f"Traitement de {employee.prenom} {employee.nom}...")
                
                # ----- 2. CHARGER L'ENTREPRISE -----
                company = session.query(CompanyInfo).first()
                if not company:
                    raise ValueError("Aucune entreprise configurée")
                
                # ----- 3. CHARGER LES SALAIRES -----
                self.progress_updated.emit(30, "Récupération des salaires...")
                
                salaries = session.query(Payslip)\
                    .filter(Payslip.employee_id == employee_id)\
                    .order_by(Payslip.period_year.desc(), Payslip.period_month.desc())\
                    .limit(12)\
                    .all()
                
                # ----- 4. PRÉPARER LE FICHIER -----
                if not output_path:
                    desktop = Path.home() / "Desktop"
                    nom = employee.nom.replace(' ', '_') if employee.nom else 'inconnu'
                    prenom = employee.prenom.replace(' ', '_') if employee.prenom else 'inconnu'
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"AS-08_{nom}_{prenom}_{timestamp}.xlsx"
                    output_path = str(desktop / filename)
                
                self.progress_updated.emit(50, "Génération du fichier Excel...")
                
                # ----- 5. CRÉER LE CLASSEUR EXCEL -----
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "AS-08"
                
                # ===== STYLES =====
                bold_font = Font(bold=True)
                title_font = Font(bold=True, size=14)
                header_font = Font(bold=True, size=11)
                normal_font = Font(size=10)
                
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # ============================================
                # 1. EN-TÊTE (lignes 1-4)
                # ============================================
                ws.merge_cells('A1:E1')
                ws['A1'] = "IMP.CNAS 03-2021 - AS.08"
                ws['A1'].font = Font(bold=True, size=16)
                ws['A1'].alignment = center_align
                
                ws.merge_cells('A2:E2')
                ws['A2'] = "ATTESTATION DE TRAVAIL ET DE SALAIRE"
                ws['A2'].font = Font(bold=True, size=14)
                ws['A2'].alignment = center_align
                
                ws.merge_cells('A3:E3')
                ws['A3'] = "شهادة عمل وأجر"
                ws['A3'].font = Font(size=12)
                ws['A3'].alignment = center_align
                
                ws.row_dimensions[4].height = 15
                
                # ============================================
                # 2. INFORMATIONS EMPLOYÉ (lignes 5-12)
                # ============================================
                row = 5
                
                # NOM
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Nom / الاسم :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = employee.nom or ""
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # PRÉNOM
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Prénom / اللقب :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = employee.prenom or ""
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # DATE NAISSANCE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Date naissance / تاريخ الميلاد :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:D{row}')
                ws[f'C{row}'] = self._format_date(employee.date_naissance)
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # LIEU NAISSANCE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Lieu naissance / مكان الميلاد :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = getattr(employee, 'lieu_naissance', '')
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # ADRESSE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Adresse / العنوان :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = employee.adresse or ""
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # MATRICULE CNAS
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Matricule CNAS / رقم التسجيل :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = getattr(employee, 'numero_secu', '')
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 2
                
                # ============================================
                # 3. EMPLOI (lignes 14-17)
                # ============================================
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "EMPLOI ACTUEL / الوظيفة الحالية"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].alignment = center_align
                ws[f'A{row}'].fill = header_fill
                row += 1
                
                # DATE EMBAUCHE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Date d'embauche / تاريخ التوظيف :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = self._format_date(employee.date_embauche)
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # FONCTION
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Fonction / المنصب :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = employee.poste or ""
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # CATÉGORIE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Catégorie / الفئة :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = getattr(employee, 'categorie', '')
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 2
                
                # ============================================
                # 4. ENTREPRISE (lignes 19-24)
                # ============================================
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "EMPLOYEUR / المستخدم"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].alignment = center_align
                ws[f'A{row}'].fill = header_fill
                row += 1
                
                # RAISON SOCIALE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Raison sociale / الاسم :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = company.nom or ""
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # ADRESSE
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Adresse / العنوان :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                adresse = f"{company.adresse or ''}"
                if company.ville:
                    adresse += f", {company.ville}"
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = adresse
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # N° EMPLOYEUR
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "N° Employeur / رقم المستخدم :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = getattr(company, 'numero_employeur', '')
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 1
                
                # RC / NIF / NIS
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "RC / NIF / NIS :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = left_align
                
                rc_nif_nis = f"{company.rc or ''} / {company.nif or ''} / {company.nis or ''}"
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = rc_nif_nis
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = left_align
                row += 2
                
                # ============================================
                # 5. TABLEAU DES SALAIRES (lignes 26-41)
                # ============================================
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "SALAIRES DES 12 DERNIERS MOIS / أجور الـ 12 شهرا الأخيرة"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].alignment = center_align
                ws[f'A{row}'].fill = header_fill
                row += 1
                
                # En-têtes français
                headers_fr = [
                    "Mois et année\nde référence",
                    "Jours\ntravaillés",
                    "Motif\nabsences",
                    "Salaire soumis\nà cotisations",
                    "Cotisation\n(part ouvrière) 9%"
                ]
                
                for col, header in enumerate(headers_fr, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.fill = header_fill
                    cell.border = thin_border
                row += 1
                
                # En-têtes arabes
                headers_ar = [
                    "الشهر والسنة\nاللذان يؤخذان كمرجع",
                    "عدد الأيام\nالمعمول فيها",
                    "سبب\nالغيابات",
                    "الأجر الخاضع\nللاشتراكات",
                    "مبلغ الإشتراك\n(حصة العامل)"
                ]
                
                for col, header in enumerate(headers_ar, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.fill = header_fill
                    cell.border = thin_border
                row += 1
                
                # Données des 12 mois
                mois_fr = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                          'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
                
                from dateutil.relativedelta import relativedelta
                today = datetime.now()
                
                salary_dict = {f"{p.period_year}-{p.period_month:02d}": p for p in salaries}
                
                for i in range(12):
                    date_periode = today - relativedelta(months=i)
                    annee = date_periode.year
                    mois = date_periode.month
                    key = f"{annee}-{mois:02d}"
                    pay = salary_dict.get(key)
                    
                    # Mois et année
                    ws.cell(row=row, column=1).value = f"{mois_fr[mois-1]} {annee}"
                    ws.cell(row=row, column=1).alignment = left_align
                    ws.cell(row=row, column=1).border = thin_border
                    
                    # Jours travaillés
                    if pay:
                        jours = pay.actual_worked_days or pay.working_days or 26
                        ws.cell(row=row, column=2).value = jours
                    else:
                        ws.cell(row=row, column=2).value = "-"
                    ws.cell(row=row, column=2).alignment = center_align
                    ws.cell(row=row, column=2).border = thin_border
                    
                    # Motif absence
                    motif = ""
                    if pay and hasattr(pay, 'absence_reason') and pay.absence_reason:
                        motif = pay.absence_reason
                    ws.cell(row=row, column=3).value = motif
                    ws.cell(row=row, column=3).alignment = left_align
                    ws.cell(row=row, column=3).border = thin_border
                    
                    # Salaire
                    if pay and pay.gross_salary:
                        salaire = self._to_float(pay.gross_salary)
                        ws.cell(row=row, column=4).value = salaire
                        ws.cell(row=row, column=4).number_format = '#,##0 "DA"'
                    ws.cell(row=row, column=4).alignment = right_align
                    ws.cell(row=row, column=4).border = thin_border
                    
                    # Cotisation
                    if pay and pay.gross_salary:
                        salaire = self._to_float(pay.gross_salary)
                        cotisation = salaire * 0.09
                        ws.cell(row=row, column=5).value = cotisation
                        ws.cell(row=row, column=5).number_format = '#,##0 "DA"'
                    ws.cell(row=row, column=5).alignment = right_align
                    ws.cell(row=row, column=5).border = thin_border
                    
                    row += 1
                
                row += 1
                
                # ============================================
                # 6. VOLUME HORAIRE
                # ============================================
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Volume horaire journalier :"
                ws[f'A{row}'].font = bold_font
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = "8 heures (08:00-12:00 / 13:00-17:00)"
                row += 1
                
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "الحجم الساعي اليومي :"
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].alignment = right_align
                
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = "8 ساعات"
                ws[f'C{row}'].alignment = right_align
                row += 2
                
                # ============================================
                # 7. SIGNATURE
                # ============================================
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "Cachet de l'employeur,"
                ws[f'A{row}'].font = bold_font
                row += 1
                
                ville = company.ville or "...................."
                date_jour = datetime.now().strftime('%d/%m/%Y')
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = f"Fait à {ville}, le {date_jour}"
                row += 1
                
                nom_directeur = getattr(company, 'nom_directeur', 'Le Gérant')
                fonction = getattr(company, 'fonction_directeur', 'Gérant')
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = f"Nom, prénom et qualité du signataire : {nom_directeur}, {fonction}"
                ws[f'A{row}'].font = bold_font
                row += 2
                
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Signature :"
                ws.merge_cells(f'C{row}:E{row}')
                ws[f'C{row}'] = "_________________________"
                row += 2
                
                # ============================================
                # 8. MENTIONS LÉGALES (français)
                # ============================================
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "(1) indiquer les salaires tels qu'ils figurent sur les fiches de paie correspondantes."
                ws[f'A{row}'].font = Font(size=9)
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● au mois précédant l'arrêt de travail, en cas de maladie,"
                ws[f'A{row}'].font = Font(size=9)
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● aux 09 mois précédant la date d'accouchement, en cas de maternité,"
                ws[f'A{row}'].font = Font(size=9)
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● aux 12 mois précédant l'arrêt de travail, en cas d'invalidité,"
                ws[f'A{row}'].font = Font(size=9)
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● aux 12 mois précédant l'accident de travail ou le décès."
                ws[f'A{row}'].font = Font(size=9)
                row += 1
                
                # IMPORTANT en rouge
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "IMPORTANT : La loi punit quiconque se rend coupable de fraude ou de fausse déclaration."
                ws[f'A{row}'].font = Font(bold=True, color="FF0000")
                row += 2
                
                # ============================================
                # 9. MENTIONS LÉGALES (arabe)
                # ============================================
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "(1) أذكر الأجور كما هي مبينة في بطاقة الأجر الموافقة لـ :"
                ws[f'A{row}'].font = Font(size=9)
                ws[f'A{row}'].alignment = right_align
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● الشهر الذي يسبق التوقف عن العمل في حالة مرض."
                ws[f'A{row}'].font = Font(size=9)
                ws[f'A{row}'].alignment = right_align
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● التسعة (09) أشهر التي تسبق تاريخ الولادة في حالة أومة."
                ws[f'A{row}'].font = Font(size=9)
                ws[f'A{row}'].alignment = right_align
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● الإثني عشرة (12) شهرا التي تسبق التوقف عن العمل في حالة عجز."
                ws[f'A{row}'].font = Font(size=9)
                ws[f'A{row}'].alignment = right_align
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "● الإثني عشرة (12) شهرا التي تسبق حادث عمل أو وفاة."
                ws[f'A{row}'].font = Font(size=9)
                ws[f'A{row}'].alignment = right_align
                row += 1
                
                # IMPORTANT en rouge (arabe)
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "هـــام : كل شخص يقوم بتزوير أو يدلي بتصريحات غير صحيحة يعاقب من طرف القانون."
                ws[f'A{row}'].font = Font(bold=True, color="FF0000")
                ws[f'A{row}'].alignment = right_align
                
                # ===== AJUSTER LES LARGEURS DE COLONNES =====
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                
                # ===== HAUTEURS DE LIGNES =====
                for r in range(1, row + 5):
                    if r in [1, 2, 3]:  # En-tête
                        ws.row_dimensions[r].height = 25
                    elif ws.cell(row=r, column=1).value and "SALAIRES" in str(ws.cell(row=r, column=1).value):
                        ws.row_dimensions[r].height = 30
                    elif ws.cell(row=r, column=1).value and "IMPORTANT" in str(ws.cell(row=r, column=1).value):
                        ws.row_dimensions[r].height = 25
                    else:
                        ws.row_dimensions[r].height = 20
                
                # ----- SAUVEGARDER -----
                wb.save(output_path)
                
                self.progress_updated.emit(100, "✅ Attestation Excel générée avec succès!")
                self.generation_finished.emit(output_path, True)
                
                return output_path
                
        except Exception as e:
            self.error_occurred.emit(str(e))
            self.generation_finished.emit("", False)
            raise